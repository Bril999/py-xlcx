import openpyxl as op

filename = 'profkom.xlsx'

choice = input('Введите действие(добавить баллы; вывести одного; вывести всех): ')
wb = op.load_workbook(filename, data_only=True)
ws = wb['Лист1']
sheet = wb.active
max_rows = sheet.max_row

if choice == 'добавить баллы':
    stud = input('Введите номер студентечского билета: ')
    mark = int(input('Введите количество баллов: '))
    for i in range(1, max_rows + 1):
        stud_check = sheet.cell(row = i, column = 4).value
        new_mark = int(sheet.cell(row = i, column = 6).value)
        if stud_check == stud:
            value = mark + new_mark
            cell = ws.cell(row=i, column=6)
            cell.value = value
        else:
            continue
if choice == 'вывести одного':
    stud = input('Введите номер студентечского билета: ')
    for i in range(1, max_rows + 1):
        stud_check = sheet.cell(row = i, column = 4).value
        if stud_check == stud:
            for j in range(1,7):
                print(sheet.cell(row = i, column = j).value, end=' ')

wb.save(filename)
wb.close